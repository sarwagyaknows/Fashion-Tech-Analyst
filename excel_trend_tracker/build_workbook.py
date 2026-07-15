import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color="1F2937")
NOTE_FONT = Font(name=FONT_NAME, italic=True, size=9, color="6B7280")
INPUT_FILL = PatternFill("solid", fgColor="FFF9C4")
BODY_FONT = Font(name=FONT_NAME, size=10)
thin = Side(style="thin", color="D1D5DB")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

with open("raw_inventory.csv", newline="", encoding="utf-8") as f:
    reader = csv.reader(f)
    header = next(reader)
    data = list(reader)

wb = Workbook()

# ---------------- Sheet 1: Raw_Inventory ----------------
ws1 = wb.active
ws1.title = "Raw_Inventory"
ws1["A1"] = "Raw Seasonal Inventory Data"
ws1["A1"].font = TITLE_FONT
ws1["A2"] = "Synthetic dataset generated to model real seasonal sell-through patterns (no live scrape source available)."
ws1["A2"].font = NOTE_FONT

for col_idx, col_name in enumerate(header, start=1):
    cell = ws1.cell(row=4, column=col_idx, value=col_name)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = BORDER

for r, row in enumerate(data, start=5):
    for c, val in enumerate(row, start=1):
        cell = ws1.cell(row=r, column=c)
        if header[c - 1] in ("Units_Sold", "Units_In_Stock", "Discount_Pct"):
            cell.value = int(val)
        elif header[c - 1] == "Price_INR":
            cell.value = float(val)
            cell.number_format = "#,##0.00"
        else:
            cell.value = val
        cell.font = BODY_FONT
        cell.border = BORDER

for i, col_name in enumerate(header, start=1):
    ws1.column_dimensions[get_column_letter(i)].width = max(12, len(col_name) + 3)

last_row = 4 + len(data)  # header at row 4, data starts row 5

# ---------------- Sheet 2: Seasonal_Summary (pivot-style, formula driven) ----------------
ws2 = wb.create_sheet("Seasonal_Summary")
ws2["A1"] = "Seasonal Demand Summary (Category x Season)"
ws2["A1"].font = TITLE_FONT
ws2["A2"] = "Built with SUMIFS / AVERAGEIFS against Raw_Inventory (recalculates automatically if raw data changes)."
ws2["A2"].font = NOTE_FONT

categories = sorted(set(row[1] for row in data))
seasons = ["Summer", "Monsoon", "Winter", "Spring"]

headers2 = ["Category", "Season", "Total_Units_Sold", "Avg_Units_In_Stock",
            "Avg_Price_INR", "Avg_Discount_Pct", "Stock_Cover_Ratio"]
for c, h in enumerate(headers2, start=1):
    cell = ws2.cell(row=4, column=c, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = BORDER

row_ptr = 5
for cat in categories:
    for season in seasons:
        ws2.cell(row=row_ptr, column=1, value=cat).border = BORDER
        ws2.cell(row=row_ptr, column=2, value=season).border = BORDER

        # Total units sold: SUMIFS(Raw range, category criteria, season criteria)
        f_sold = (f"=SUMIFS(Raw_Inventory!E5:E{last_row},"
                  f"Raw_Inventory!B5:B{last_row},A{row_ptr},"
                  f"Raw_Inventory!C5:C{last_row},B{row_ptr})")
        c3 = ws2.cell(row=row_ptr, column=3, value=f_sold)
        c3.border = BORDER

        f_stock = (f"=IFERROR(AVERAGEIFS(Raw_Inventory!F5:F{last_row},"
                   f"Raw_Inventory!B5:B{last_row},A{row_ptr},"
                   f"Raw_Inventory!C5:C{last_row},B{row_ptr}),0)")
        c4 = ws2.cell(row=row_ptr, column=4, value=f_stock)
        c4.number_format = "#,##0"
        c4.border = BORDER

        f_price = (f"=IFERROR(AVERAGEIFS(Raw_Inventory!G5:G{last_row},"
                   f"Raw_Inventory!B5:B{last_row},A{row_ptr},"
                   f"Raw_Inventory!C5:C{last_row},B{row_ptr}),0)")
        c5 = ws2.cell(row=row_ptr, column=5, value=f_price)
        c5.number_format = "#,##0.00"
        c5.border = BORDER

        f_disc = (f"=IFERROR(AVERAGEIFS(Raw_Inventory!H5:H{last_row},"
                  f"Raw_Inventory!B5:B{last_row},A{row_ptr},"
                  f"Raw_Inventory!C5:C{last_row},B{row_ptr}),0)")
        c6 = ws2.cell(row=row_ptr, column=6, value=f_disc)
        c6.number_format = "0.0"
        c6.border = BORDER

        # Stock cover ratio = avg stock / (total sold / months in that season), guarded for /0
        f_cover = f"=IFERROR(D{row_ptr}/(C{row_ptr}/4),0)"
        c7 = ws2.cell(row=row_ptr, column=7, value=f_cover)
        c7.number_format = "0.00"
        c7.border = BORDER

        for c in range(1, 8):
            ws2.cell(row=row_ptr, column=c).font = BODY_FONT
        row_ptr += 1

for i, w in enumerate([14, 12, 18, 20, 16, 18, 18], start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2[f"A{row_ptr+1}"] = "Reading it: high Total_Units_Sold + low Stock_Cover_Ratio = fast-fashion demand spike, restock priority."
ws2[f"A{row_ptr+1}"].font = NOTE_FONT

# ---------------- Sheet 3: SKU_Lookup ----------------
ws3 = wb.create_sheet("SKU_Lookup")
ws3["A1"] = "SKU Lookup Tool"
ws3["A1"].font = TITLE_FONT
ws3["A2"] = ("Type a SKU in the yellow cell. Uses INDEX/MATCH here (this sandbox's LibreOffice "
             "engine can't evaluate XLOOKUP), but this is functionally identical to XLOOKUP and "
             "the formula bar shows the underlying logic clearly.")
ws3["A2"].font = NOTE_FONT
ws3.column_dimensions["A"].width = 20
ws3.column_dimensions["B"].width = 40

ws3["A4"] = "Enter SKU:"
ws3["A4"].font = Font(name=FONT_NAME, bold=True, size=10)
ws3["B4"] = data[0][0]  # example SKU pre-filled
ws3["B4"].fill = INPUT_FILL
ws3["B4"].font = BODY_FONT
ws3["B4"].border = BORDER

labels_formulas = [
    ("Category", 2),
    ("Season", 3),
    ("Month", 4),
    ("Units_Sold", 5),
    ("Units_In_Stock", 6),
    ("Price_INR", 7),
    ("Discount_Pct", 8),
]
for i, (label, col_num) in enumerate(labels_formulas, start=6):
    ws3.cell(row=i, column=1, value=label).font = Font(name=FONT_NAME, bold=True, size=10)
    formula = (f"=IFERROR(INDEX(Raw_Inventory!{get_column_letter(col_num)}5:"
               f"{get_column_letter(col_num)}{last_row},"
               f"MATCH($B$4,Raw_Inventory!A5:A{last_row},0)),\"Not found\")")
    cell = ws3.cell(row=i, column=2, value=formula)
    cell.font = BODY_FONT
    cell.border = BORDER

for r in range(4, 13):
    ws3.cell(row=r, column=1).border = BORDER
    ws3.cell(row=r, column=2).border = BORDER

wb.save("Fashion_Trend_Tracker.xlsx")
print("Saved Fashion_Trend_Tracker.xlsx")
